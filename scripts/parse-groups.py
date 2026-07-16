import openpyxl
import json
import os

xlsx_path = '/home/massimo/Scrivania/Siti/Morgana-Orum-v2/public/assets/documents/gruppi 26-27.xlsx'
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
sheet = wb.active

deptMapping = {
    'GIURISPRUDENZA': 'Dipartimento di Giurisprudenza',
    'DICAM': 'Dipartimento di Civiltà Antiche e Moderne (DICAM)',
    'ECONOMIA': 'Dipartimento di Economia',
    'INGEGNERIA': 'Dipartimento di Ingegneria',
    'DIMED': 'Dipartimento di Medicina Clinica e Sperimentale (DIMED)',
    'PATOLOGIAUMANADETEV': "Dipartimento di Patologia Umana dell'Adulto e dell'Età Evolutiva \"Gaetano Barresi\"",
    'BIOMORF': 'Dipartimento di Scienze Biomediche, Odontoiatriche e delle Immagini Morfologiche e Funzionali (BIOMORF)',
    'CHIBIOFARAM': 'Dipartimento di Scienze Chimiche, Biologiche, Farmaceutiche e Ambientali (CHIBIOFARAM)',
    'COSPECS': 'Dipartimento di Scienze Cognitive, Psicologiche, Pedagogiche e degli Studi Culturali (COSPECS)',
    'MIFT': 'Dipartimento di Scienze Matematiche e Informatiche, Scienze Fisiche e Scienze della Terra (MIFT)',
    'SCIPOG': 'Dipartimento di Scienze Politiche e Giuridiche (SCIPOG)',
    'VETERINARIA': 'Dipartimento di Scienze Veterinarie'
}

current_dept = None
dept_order = 0
raw_records = []

for r in range(1, sheet.max_row + 1):
    # Read cells A to E
    code_val = sheet.cell(r, 1).value
    course_val = sheet.cell(r, 2).value
    link_val = sheet.cell(r, 3).value
    sede_val = sheet.cell(r, 4).value
    note_val = sheet.cell(r, 5).value

    # Trim strings
    code = str(code_val).strip() if code_val is not None else ""
    course = str(course_val).strip() if course_val is not None else ""
    link = str(link_val).strip() if link_val is not None else ""
    sede = str(sede_val).strip() if sede_val is not None else ""
    note = str(note_val).strip() if note_val is not None else ""

    # Skip header row
    if code == 'CODICE' and course == 'CORSO' and link == 'LINK':
        continue

    # Check if department header
    if code in deptMapping and not course and not link:
        current_dept = deptMapping[code]
        dept_order = 0
        continue

    # Check if row is yellow
    is_yellow = False
    for c in range(1, 6):
        cell = sheet.cell(r, c)
        fill = cell.fill
        if fill and fill.fill_type:
            color = fill.start_color
            if color:
                if color.type == 'theme' and color.value == 7:
                    is_yellow = True
                elif color.type == 'rgb' and color.value in ('FFFFFFFF00', 'FFFF00', 'FFFFE0'):
                    is_yellow = True

    # Skip if yellow
    if is_yellow:
        continue

    # Skip if empty (no course and no link)
    if not course and not link:
        continue

    # Skip if link is missing
    if not link:
        continue

    # We must have a department
    if not current_dept:
        continue

    # Increment order for sorting within department
    dept_order += 1

    raw_records.append({
        'rowNumber': r,
        'code': code,
        'course': course,
        'link': link,
        'sede': sede,
        'department': current_dept,
        'order': dept_order,
        'note': note
    })

# Now, detect duplicate course names in the same department with multiple DIFFERENT locations
# We group by (department, course)
from collections import defaultdict
grouped = defaultdict(list)
for rec in raw_records:
    key = (rec['department'], rec['course'])
    grouped[key].append(rec)

final_records = []
for key, group in grouped.items():
    # Count unique locations for this course
    unique_sedi = set(rec['sede'] for rec in group if rec['sede'])
    has_multiple_locations = len(unique_sedi) > 1
    
    for rec in group:
        # Base name: "[Code] [Course]"
        base_name = f"{rec['code']} {rec['course']}".strip() if rec['code'] else rec['course']
        
        # If the course actually has multiple locations in the list, append it
        if has_multiple_locations and rec['sede']:
            title_sede = rec['sede'].title() # e.g. "Messina" or "Siracusa"
            name = f"{base_name} ({title_sede})"
        else:
            name = base_name
            
        final_records.append({
            'name': name,
            'link': rec['link'],
            'category': 'ACADEMIC',
            'department': rec['department'],
            'order': rec['order'],
            'semester': '2026/2027',
            'isGeneral': False
        })

# Sort final records by department, then by order
final_records.sort(key=lambda x: (x['department'], x['order']))

output_path = '/home/massimo/.gemini/antigravity-ide/brain/05abbf9d-8321-4020-9f44-5c08defabdeb/scratch/cleaned_groups.json'
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(final_records, f, indent=2, ensure_ascii=False)

print(f"Refined duplicate logic. Parsed {len(final_records)} valid groups successfully!")
print(f"JSON saved to {output_path}")
